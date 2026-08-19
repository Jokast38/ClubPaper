"""Excel (.xlsx) import."""
import io
from openpyxl import load_workbook


HEADER_MAP = {
    "prenom": "first_name", "prénom": "first_name", "firstname": "first_name", "first_name": "first_name",
    "nom": "last_name", "lastname": "last_name", "last_name": "last_name",
    "email": "email", "e-mail": "email", "mail": "email",
    "telephone": "phone", "téléphone": "phone", "tel": "phone", "phone": "phone",
    "equipe": "team", "équipe": "team", "team": "team",
    "date_naissance": "birth_date", "naissance": "birth_date", "birth_date": "birth_date",
    "parent": "parent_name", "parent_name": "parent_name",
    "parent_email": "parent_email",
}


def parse_xlsx(content: bytes):
    """Return list of dicts (member fields) parsed from an xlsx file."""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    mapped = [HEADER_MAP.get(h, None) for h in headers]
    out = []
    for row in rows[1:]:
        rec = {}
        for i, cell in enumerate(row):
            k = mapped[i] if i < len(mapped) else None
            if not k:
                continue
            v = "" if cell is None else str(cell).strip()
            rec[k] = v
        if rec.get("first_name") and rec.get("last_name"):
            out.append(rec)
    return out
